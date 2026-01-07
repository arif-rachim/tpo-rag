#!/usr/bin/env python3

import json
import os
import re
import sys
from datetime import datetime
from typing import Dict, Any, Optional, Set

from openpyxl import load_workbook
from openpyxl.styles.colors import RGB

# Constants
OUTPUT_DIR = "output"
CELL_REF_PATTERN = re.compile(r'([A-Za-z]+[0-9]+(?::[A-Za-z]+[0-9]+)?)')
NAMED_RANGE_PATTERN = re.compile(r'[A-Za-z][A-Za-z0-9_.]*')

class ExcelJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder to handle Excel-specific types."""
    def default(self, obj):
        if isinstance(obj, RGB):
            # Convert RGB object to hex string or return None if no color
            if hasattr(obj, "rgb"):
                return obj.rgb
            elif hasattr(obj, "index"):
                # Handle theme colors
                return f"theme_{obj.index}"
            else:
                return None
        elif isinstance(obj, datetime):
            # Convert datetime objects to ISO format string
            return obj.isoformat()
        return super().default(obj)


def _get_column_letter(col_num: int) -> str:
    """Convert column number to letter (1 = A, 2 = B, etc.)."""
    result = ""
    while col_num:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result

def _log_warning(message: str, error: Exception) -> None:
    """Log a warning message with error details."""
    print(f"Warning: {message}: {str(error)}")

def _extract_cell_dependencies(formula: str) -> Set[str]:
    """Extract cell references from a formula."""
    if not formula:
        return set()
    
    try:
        # Find all cell references in the formula
        cell_refs = set(CELL_REF_PATTERN.findall(formula))
        # Remove any string literals that might look like cell references
        return {ref for ref in cell_refs if not ref.startswith('"') and not ref.startswith("'")}
    except Exception as e:
        _log_warning("Could not extract dependencies from formula", e)
        return set()

def _process_array_formula(cell: Any) -> Dict[str, Any] | None:
    """Process array formula and its evaluation context."""
    try:
        if not (isinstance(cell.value, str) and cell.value.startswith("{")):
            return None
        parts = cell.value[1:-1].split(";")
        return {
            "type": "array_formula",
            "formula": cell.value[1:-1],
            "range": cell.coordinate,
            "calculated_value": cell.internal_value,
            "dimensions": {"rows": len(parts), "columns": len(parts[0].split(","))}
        }
    except Exception as e:
        _log_warning("Failed to process array formula", e)
        return None

def _extract_cell_style(cell: Any) -> Dict[str, Any]:
    """Extract style information from a cell."""
    return {
        "font": {
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "color": cell.font.color.rgb if cell.font.color else None
        },
        "fill": {
            "background": getattr(getattr(cell.fill, "start_color", None), "rgb", None)
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical
        }
    }

def excel_to_json(file_path: str, sample_size: Optional[int] = None) -> Optional[Dict[str, Any]]:
    """
    Convert Excel file to JSON structure with optional row sampling.
    
    Args:
        file_path: Path to the Excel file
        sample_size: Optional maximum number of rows to process per sheet
        
    Returns:
        Dict containing the spreadsheet structure or None if an error occurs
    """
    try:
        # Validate file existence
        if not os.path.exists(file_path):
            print(f"Error: File not found: {file_path}", "error")
            return None

        # Validate file extension
        if not any(file_path.lower().endswith(ext) for ext in [".xlsx", ".xls", ".xlsm"]):
            print(f"Error: Unsupported file format. File must be .xlsx, .xls, or .xlsm", "error")
            return None

        try:
            wb = load_workbook(filename=file_path, data_only=False)
        except Exception as e:
            print(f"Error: Failed to load Excel file: {str(e)}", "error")
            return None
        
        spreadsheet_dict = {
            "fileName": os.path.basename(file_path),
            "sheets": []
        }

        for sheet_name in wb.sheetnames:
            try:
                sheet = wb[sheet_name]
                sheet_data = {
                    "sheetTitle": sheet_name,
                    "maxRow": sheet.max_row,
                    "maxColumn": sheet.max_column,
                    "cells": {}
                }

                # Determine how many rows we should actually iterate
                row_limit = sheet.max_row if sample_size is None else min(sample_size, sheet.max_row)
                

                # Process cells
                for row_index, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                    if row_index > row_limit:
                        break

                    for cell in row:
                        try:
                            cell_coord = cell.coordinate
                            cell_data = {"value": cell.value}

                            # Handle array formulas
                            array_formula = _process_array_formula(cell)
                            if array_formula:
                                cell_data["array_formula"] = array_formula

                            # Handle regular formulas
                            if cell.data_type == "f":
                                cell_data.update({
                                    "formula": cell.value,
                                    "calculated_value": cell.internal_value,
                                    "dependencies": list(_extract_cell_dependencies(cell.value))
                                })

                            # Add style information
                            cell_data["style"] = _extract_cell_style(cell)

                            # Only add non-empty cells to reduce output size
                            if cell.value is not None:
                                sheet_data["cells"][cell_coord] = cell_data

                        except Exception as e:
                            _log_warning(f"Failed to process cell {cell_coord} in sheet {sheet_name}", e)
                            continue

                spreadsheet_dict["sheets"].append(sheet_data)

            except Exception as e:
                _log_warning(f"Failed to process sheet {sheet_name}", e)
                continue

        if not spreadsheet_dict["sheets"]:
            print("Error: No valid sheets were processed", "error")
            return None

        return spreadsheet_dict

    except Exception as e:
        print(f"Error converting Excel file: {str(e)}", "error")
        return None


if __name__ == "__main__":
    # Example usage
    if len(sys.argv) > 1:
        try:
            result = excel_to_json(sys.argv[1], sample_size=int(sys.argv[2]) if len(sys.argv) > 2 else None)
            if result:
                sys.exit(0)
            else:
                sys.exit(1)
        except Exception as e:
            print(f"Error: {str(e)}", "error")
            sys.exit(1)
    else:
        sys.exit(1)